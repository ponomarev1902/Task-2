import plistlib
import argparse
import xml.etree.ElementTree as ET
from openpyxl import load_workbook


def read_pls(file_name):
    """Read data from .plist file and return it as dictionary"""
    with open(file_name, 'rb') as fp:
        return plistlib.load(fp)


def read_xml(file_name):
    """Read data from .xml file end return the corresponding tree """
    tree = ET.parse(file_name)
    return tree.getroot()[0]


def check_xml(key, value, file_name='DominiIAP.xml', id='ProductID', store='store_desc'):
    """Compare the data in .xml and .xlsx files"""
    for itm in read_xml(file_name):
        if itm.find(id).text == key:
            if itm.find(store).text != value:
                print(f'Error in {file_name}:', itm.find(store).text)
            break
    else:
        print(f'No key {key} in {file_name}')


def check_pls(key, value, file_name='Info.plist'):
    """Compare the data in .plist and .xlsx files"""
    pls = read_pls(file_name)
    if key in pls:
        if pls[key] != value:
            print(f'Error in {file_name}:', value)
    else:
        print(f'No key {key} in {file_name}')


def main(xls_file='DominiGames Test  Sheet.xlsx'):
    wb = load_workbook(xls_file)
    ws = wb.active

    for x, y in zip(ws["A"][1:], ws["B"][1:]):
        if x.value is not None:
            check_pls(x.value, y.value)

    for x, y in zip(ws["C"][1:], ws["D"][1:]):
        if x.value is not None:
            check_xml(x.value, y.value)


def parse():
    parser = argparse.ArgumentParser()
    parser.add_argument('--path', help='Path')
    return parser.parse_args()


if __name__ == '__main__':
    args = parse()
    if args.path:
        main(args.path)
    else:
        main()

